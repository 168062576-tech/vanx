"""
御龙军虚拟世界 - 实验报告生成器

根据用户输入的实验目标、成功标准、关心指标，生成针对性的分析报告。
不是导出原始数据，而是回答用户问题：
- 成功/失败？
- 谁买了？（画像）
- 谁没买？为什么？
- 下一步建议？

作者：御龙军
日期：2026-03-19
移植时间：2026-03-24 (v5 适配)
"""

import os
import json
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass

# 依赖本地模块
try:
    from data_exporter import DataExporter
except ImportError:
    DataExporter = None

try:
    from visual_report_generator import VisualReportGenerator
except ImportError:
    VisualReportGenerator = None

try:
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class BuyerPersona:
    """购买者画像统计"""
    age_distribution: Dict[str, int]  # 年龄段 -> 人数
    income_distribution: Dict[str, int]  # 收入段 -> 人数
    occupation_distribution: Dict[str, int]  # 职业 -> 人数
    education_distribution: Dict[str, int]  # 教育 -> 人数
    marital_distribution: Dict[str, int]  # 婚姻 -> 人数
    avg_age: float
    avg_income: float
    top_occupation: str
    top_education: str


@dataclass
class NonBuyerReason:
    """未购买原因统计"""
    price_sensitive: int  # 价格敏感
    not_needed: int  # 不需要
    has_alternative: int  # 已有替代品
    other: int  # 其他
    total: int

    @property
    def price_sensitive_pct(self) -> float:
        return self.price_sensitive / self.total * 100 if self.total > 0 else 0

    @property
    def not_needed_pct(self) -> float:
        return self.not_needed / self.total * 100 if self.total > 0 else 0

    @property
    def has_alternative_pct(self) -> float:
        return self.has_alternative / self.total * 100 if self.total > 0 else 0


class ExperimentReportGenerator:
    """实验报告生成器 - 根据模板类型生成不同报告"""

    # 模板类型 -> 报告配置
    TEMPLATE_REPORT_CONFIG = {
        'edu': {
            'persona_fields': ['age', 'education_level', 'parent_income', 'region'],
            'metrics_focus': ['升学率', '就业率', '教育满意度'],
            'chart_types': ['education_dist', 'region_dist', 'income_comparison'],
            'analysis': ['policy_impact', 'equity_analysis']
        },
        'business': {
            'persona_fields': ['age', 'income', 'occupation', 'education_level', 'marital_status'],
            'metrics_focus': ['转化率', '复购率', 'NPS', 'ROI'],
            'chart_types': ['age_dist', 'income_dist', 'occupation_pie'],
            'analysis': ['buyer_persona', 'non_buyer_reasons', 'price_sensitivity']
        },
        'health': {
            'persona_fields': ['age', 'health_score', 'disease_type', 'region'],
            'metrics_focus': ['治愈率', '满意度', '医疗成本'],
            'chart_types': ['health_dist', 'disease_pie', 'cost_comparison'],
            'analysis': ['treatment_effectiveness', 'cost_benefit']
        },
        'policy': {
            'persona_fields': ['age', 'income', 'occupation', 'region'],
            'metrics_focus': ['支持率', '影响人数', '满意度'],
            'chart_types': ['support_dist', 'impact_bar', 'region_heatmap'],
            'analysis': ['demographic_impact', 'regional_analysis']
        },
        'social': {
            'persona_fields': ['age', 'occupation', 'region', 'relationships'],
            'metrics_focus': ['参与度', '影响力', '社区活跃度'],
            'chart_types': ['age_dist', 'region_dist', 'participation_pie'],
            'analysis': ['community_impact', 'cultural_preservation']
        },
        'tech': {
            'persona_fields': ['age', 'occupation', 'education_level', 'tech_adoption'],
            'metrics_focus': ['采用率', '效率提升', '创新指数'],
            'chart_types': ['adoption_curve', 'efficiency_bar', 'innovation_dist'],
            'analysis': ['tech_adoption', 'productivity_gain']
        },
    }

    def __init__(self, output_dir: str = None):
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

        self.data_exporter = DataExporter(output_dir) if DataExporter else None
        self.visual_report = VisualReportGenerator(output_dir) if VisualReportGenerator else None

    def generate_full_report(self, experiment_id: str, results: Dict, user_config: Dict) -> Dict[str, str]:
        """
        生成完整实验报告

        Args:
            experiment_id: 实验 ID
            results: 实验结果数据（来自 experiment_runner）
            user_config: 用户配置 {
                'template_name': str,
                'key_metrics': List[str],
                'success_criteria': Dict{metric: target},
                'target_audience': {'age_range': (min, max), 'income_min': float}
            }

        Returns:
            {'excel': path, 'csv': path, 'summary': path}
        """
        print(f"\n[REPORT] 生成实验报告: {experiment_id}")

        generated_files = {}
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 1. Excel 报告（带图表）
        excel_path = self._generate_excel_report(experiment_id, results, user_config, timestamp)
        if excel_path:
            generated_files['excel'] = excel_path

        # 2. CSV 数据（购买者名单 + 月度数据）
        csv_path = self._generate_csv_data(experiment_id, results, timestamp)
        if csv_path:
            generated_files['csv'] = csv_path

        # 3. 文本摘要（浏览器可直接查看）
        summary_path = self._generate_text_summary(experiment_id, results, user_config, timestamp)
        if summary_path:
            generated_files['summary'] = summary_path

        print(f"[OK] 报告生成完成：{len(generated_files)} 个文件")
        return generated_files

    def generate_excel_from_engine(self, engine, world_name: str = 'default') -> Optional[str]:
        """
        从引擎数据直接生成 Excel 报告（用于 API 调用）

        Args:
            engine: DeepIntegrationEngine 实例
            world_name: 世界名称

        Returns:
            Excel 文件路径，或 None
        """
        if not OPENPYXL_AVAILABLE:
            return None

        try:
            wb = Workbook()
            wb.remove(wb.active)

            agents_list = list(engine.agents.values())
            alive = [a for a in agents_list if a.is_alive]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Sheet 1: 执行摘要
            ws1 = wb.create_sheet('执行摘要')
            ws1['A1'] = f"虚拟世界报告：{world_name}"
            ws1['A1'].font = Font(size=18, bold=True)
            ws1['A3'] = "生成时间"
            ws1['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")
            ws1['A4'] = "总 Agent 数"
            ws1['B4'] = len(agents_list)
            ws1['A5'] = "存活人口"
            ws1['B5'] = len(alive)
            ws1['A6'] = "已模拟月数"
            ws1['B6'] = engine.months_simulated

            # Sheet 2: Agent 明细
            ws2 = wb.create_sheet('Agent 明细')
            headers = ['ID', '姓名', '年龄', '性别', '职业', '收入', '教育', '婚姻', '健康', '幸福度', '存活']
            for col, h in enumerate(headers, 1):
                ws2.cell(row=1, column=col, value=h)
                ws2.cell(row=1, column=col).font = Font(bold=True)

            for row_idx, agent in enumerate(agents_list[:5000], 2):
                d = agent.to_dict()
                ws2.cell(row=row_idx, column=1, value=d.get('id'))
                ws2.cell(row=row_idx, column=2, value=d.get('name', ''))
                ws2.cell(row=row_idx, column=3, value=d.get('age'))
                ws2.cell(row=row_idx, column=4, value=d.get('gender'))
                ws2.cell(row=row_idx, column=5, value=d.get('occupation', ''))
                ws2.cell(row=row_idx, column=6, value=d.get('income', 0))
                ws2.cell(row=row_idx, column=7, value=d.get('education_level', ''))
                ws2.cell(row=row_idx, column=8, value=d.get('marital_status', ''))
                ws2.cell(row=row_idx, column=9, value=d.get('health_score', 0))
                ws2.cell(row=row_idx, column=10, value=d.get('happiness', 0))
                ws2.cell(row=row_idx, column=11, value='是' if agent.is_alive else '否')

            # Sheet 3: 职业分布
            ws3 = wb.create_sheet('职业分布')
            occ_stats = {}
            for a in alive:
                d = a.to_dict()
                job = d.get('occupation', d.get('job', '未知'))
                occ_stats[job] = occ_stats.get(job, 0) + 1

            ws3['A1'] = '职业'
            ws3['B1'] = '人数'
            ws3['C1'] = '占比'
            for c in 'ABC':
                ws3[f'{c}1'].font = Font(bold=True)

            total = len(alive) or 1
            for row_idx, (job, count) in enumerate(sorted(occ_stats.items(), key=lambda x: x[1], reverse=True), 2):
                ws3.cell(row=row_idx, column=1, value=job)
                ws3.cell(row=row_idx, column=2, value=count)
                ws3.cell(row=row_idx, column=3, value=f"{count/total*100:.1f}%")

            # 添加饼图
            if len(occ_stats) > 0:
                try:
                    chart = PieChart()
                    chart.title = "职业分布"
                    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(occ_stats) + 1)
                    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(occ_stats) + 1)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    chart.width = 15
                    chart.height = 10
                    ws3.add_chart(chart, "E2")
                except Exception as e:
                    print(f"  [WARN] 职业饼图失败：{e}")

            # Sheet 4: 教育分布
            ws4 = wb.create_sheet('教育分布')
            edu_stats = {}
            for a in alive:
                d = a.to_dict()
                edu = d.get('education', d.get('education_level', '未知'))
                edu_stats[edu] = edu_stats.get(edu, 0) + 1

            ws4['A1'] = '学历'
            ws4['B1'] = '人数'
            for c in 'AB':
                ws4[f'{c}1'].font = Font(bold=True)

            for row_idx, (edu, count) in enumerate(sorted(edu_stats.items(), key=lambda x: x[1], reverse=True), 2):
                ws4.cell(row=row_idx, column=1, value=edu)
                ws4.cell(row=row_idx, column=2, value=count)

            # 添加柱状图
            if len(edu_stats) > 0:
                try:
                    chart = BarChart()
                    chart.title = "教育分布"
                    data_ref = Reference(ws4, min_col=2, min_row=1, max_row=len(edu_stats) + 1)
                    cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=len(edu_stats) + 1)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    chart.width = 15
                    chart.height = 10
                    ws4.add_chart(chart, "D2")
                except Exception as e:
                    print(f"  [WARN] 教育柱状图失败：{e}")

            # Sheet 5: 经济统计
            ws5 = wb.create_sheet('经济统计')
            incomes = [a.income for a in alive if hasattr(a, 'income') and a.income > 0]
            ws5['A1'] = '统计项'
            ws5['B1'] = '值'
            ws5['A1'].font = Font(bold=True)
            ws5['B1'].font = Font(bold=True)

            stats_data = [
                ('平均收入', f"¥{sum(incomes)/len(incomes):,.0f}" if incomes else 'N/A'),
                ('最高收入', f"¥{max(incomes):,.0f}" if incomes else 'N/A'),
                ('最低收入', f"¥{min(incomes):,.0f}" if incomes else 'N/A'),
                ('收入中位数', f"¥{sorted(incomes)[len(incomes)//2]:,.0f}" if incomes else 'N/A'),
                ('就业人数', sum(1 for a in alive if a.occupation and a.occupation not in ('unemployed', '', '无业', '失业'))),
                ('失业人数', sum(1 for a in alive if not a.occupation or a.occupation in ('unemployed', '', '无业', '失业'))),
                ('已婚人数', sum(1 for a in alive if a.marital_status == 'married')),
            ]
            for row_idx, (label, val) in enumerate(stats_data, 2):
                ws5.cell(row=row_idx, column=1, value=label)
                ws5.cell(row=row_idx, column=2, value=val)

            # 保存
            filename = f"excel_report_{world_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            print(f"  [OK] Excel 报告：{filename}")
            return filepath

        except Exception as e:
            print(f"  [ERR] Excel 报告失败：{e}")
            import traceback
            traceback.print_exc()
            return None

    def generate_csv_from_engine(self, engine, world_name: str = 'default') -> Optional[str]:
        """
        从引擎数据直接生成 CSV 文件（用于 API 调用）

        Args:
            engine: DeepIntegrationEngine 实例
            world_name: 世界名称

        Returns:
            CSV 文件路径，或 None
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            agents_list = list(engine.agents.values())

            # 构建扁平化数据
            data = []
            for agent in agents_list:
                d = agent.to_dict()
                data.append({
                    'id': d.get('id'),
                    'name': d.get('name', ''),
                    'age': d.get('age'),
                    'gender': d.get('gender'),
                    'occupation': d.get('occupation', ''),
                    'income': d.get('income', 0),
                    'education_level': d.get('education_level', ''),
                    'marital_status': d.get('marital_status', ''),
                    'health_score': d.get('health_score', 0),
                    'happiness': d.get('happiness', 0),
                    'is_alive': '是' if agent.is_alive else '否',
                })

            if not self.data_exporter:
                # 无 DataExporter 时直接写
                import csv
                filename = f"csv_export_{world_name}_{timestamp}.csv"
                filepath = os.path.join(self.output_dir, filename)
                if data:
                    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                        writer = csv.DictWriter(f, fieldnames=data[0].keys())
                        writer.writeheader()
                        writer.writerows(data)
                print(f"  [OK] CSV 导出：{filename}")
                return filepath
            else:
                filename = f"csv_export_{world_name}_{timestamp}.csv"
                return self.data_exporter.export_to_csv(data, filename)

        except Exception as e:
            print(f"  [ERR] CSV 导出失败：{e}")
            import traceback
            traceback.print_exc()
            return None

    def _generate_excel_report(self, experiment_id: str, results: Dict, user_config: Dict, timestamp: str) -> Optional[str]:
        """生成带图表的 Excel 报告"""
        if not OPENPYXL_AVAILABLE:
            print("[WARN] openpyxl 未安装，跳过 Excel 报告")
            return None

        try:
            wb = Workbook()
            wb.remove(wb.active)

            # Sheet 1: 执行摘要
            ws_summary = wb.create_sheet('执行摘要')
            self._add_executive_summary(ws_summary, experiment_id, results, user_config)

            # Sheet 2: 关键指标达成
            ws_metrics = wb.create_sheet('关键指标')
            self._add_key_metrics(ws_metrics, results, user_config)

            # Sheet 3: 购买者画像
            ws_persona = wb.create_sheet('购买者画像')
            persona = self._analyze_buyer_persona(results)
            self._add_buyer_persona(ws_persona, persona)

            # Sheet 4: 未购买原因
            ws_non_buyer = wb.create_sheet('未购买分析')
            non_buyer = self._analyze_non_buyer_reasons(results)
            self._add_non_buyer_analysis(ws_non_buyer, non_buyer)

            # Sheet 5: 建议与下一步
            ws_recommend = wb.create_sheet('改进建议')
            self._add_recommendations(ws_recommend, results, user_config, persona, non_buyer)

            # 保存
            filename = f"report_{experiment_id}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)

            print(f"  [OK] Excel 报告：{filename}")
            return filepath

        except Exception as e:
            print(f"  [ERR] Excel 报告失败：{e}")
            return None

    def _add_executive_summary(self, ws, experiment_id: str, results: Dict, user_config: Dict):
        """添加执行摘要工作表"""
        from openpyxl.styles import Font, Alignment, PatternFill

        ws['A1'] = f"实验报告：{user_config.get('template_name', experiment_id)}"
        ws['A1'].font = Font(size=18, bold=True)

        row = 3
        ws[f'A{row}'] = "实验 ID"
        ws[f'B{row}'] = experiment_id
        row += 1

        ws[f'A{row}'] = "生成时间"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        row += 2

        ws[f'A{row}'] = "实验目标"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = user_config.get('description', '未提供')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 2

        ws[f'A{row}'] = "成功标准"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for metric, target in user_config.get('success_criteria', {}).items():
            ws[f'A{row}'] = f"  • {metric}: {target*100:.0f}%"
            row += 1
        row += 1

        ws[f'A{row}'] = "实验结论"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        success = self._evaluate_success(results, user_config)
        conclusion = "✅ 成功" if success['overall_success'] else "❌ 未达目标"
        ws[f'A{row}'] = conclusion
        ws[f'A{row}'].font = Font(size=16, bold=True)
        if success['overall_success']:
            ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws[f'A{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 2

        ws[f'A{row}'] = "关键发现"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for i, finding in enumerate(success.get('key_findings', [])[:3], 1):
            ws[f'A{row}'] = f"{i}. {finding}"
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

    def _add_key_metrics(self, ws, results: Dict, user_config: Dict):
        """添加关键指标达成工作表（带柱状图）"""
        ws['A1'] = "指标"
        ws['B1'] = "实际值"
        ws['C1'] = "目标值"
        ws['D1'] = "达成率"
        ws['E1'] = "状态"

        for col in 'ABCDE':
            ws[f'{col}1'].font = Font(bold=True)

        row = 2
        metrics_data = results.get('metrics', {})
        success_criteria = user_config.get('success_criteria', {})

        for metric, target in success_criteria.items():
            actual = metrics_data.get(metric, 0)
            achievement = actual / target if target > 0 else 0
            status = "✓" if actual >= target else "✗"

            ws[f'A{row}'] = metric
            ws[f'B{row}'] = f"{actual*100:.1f}%"
            ws[f'C{row}'] = f"{target*100:.1f}%"
            ws[f'D{row}'] = f"{achievement*100:.1f}%"
            ws[f'E{row}'] = status

            if actual >= target:
                ws[f'E{row}'].font = Font(color="006100")
            else:
                ws[f'E{row}'].font = Font(color="9C0006")

            row += 1

        if row > 2:
            try:
                chart = BarChart()
                chart.title = "指标达成情况"
                chart.type = "col"
                data_ref = Reference(ws, min_col=2, min_row=1, max_row=row-1)
                cats_ref = Reference(ws, min_col=1, min_row=2, max_row=row-1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.shape = 4
                chart.width = 15
                chart.height = 10
                ws.add_chart(chart, "G2")
            except Exception as e:
                print(f"    [WARN] 图表生成失败：{e}")

    def _analyze_buyer_persona(self, results: Dict) -> BuyerPersona:
        """分析购买者画像"""
        buyers = results.get('buyers', [])

        age_dist = {}
        income_dist = {}
        occupation_dist = {}
        education_dist = {}
        marital_dist = {}

        total_age = 0
        total_income = 0

        for buyer in buyers:
            age = buyer.get('age', 0)
            total_age += age
            age_bucket = f"{age//10*10}-{age//10*10+9}岁"
            age_dist[age_bucket] = age_dist.get(age_bucket, 0) + 1

            income = buyer.get('income', 0)
            total_income += income
            if income < 5000:
                bucket = "<5k"
            elif income < 10000:
                bucket = "5k-1w"
            elif income < 20000:
                bucket = "1w-2w"
            elif income < 50000:
                bucket = "2w-5w"
            else:
                bucket = "5w+"
            income_dist[bucket] = income_dist.get(bucket, 0) + 1

            occupation = buyer.get('occupation', '未知')
            occupation_dist[occupation] = occupation_dist.get(occupation, 0) + 1

            education = buyer.get('education_level', '未知')
            education_dist[education] = education_dist.get(education, 0) + 1

            marital = buyer.get('marital_status', '未知')
            marital_dist[marital] = marital_dist.get(marital, 0) + 1

        top_occupation = max(occupation_dist, key=occupation_dist.get) if occupation_dist else "未知"
        top_education = max(education_dist, key=education_dist.get) if education_dist else "未知"

        return BuyerPersona(
            age_distribution=age_dist,
            income_distribution=income_dist,
            occupation_distribution=occupation_dist,
            education_distribution=education_dist,
            marital_distribution=marital_dist,
            avg_age=total_age / len(buyers) if buyers else 0,
            avg_income=total_income / len(buyers) if buyers else 0,
            top_occupation=top_occupation,
            top_education=top_education
        )

    def _add_buyer_persona(self, ws, persona: BuyerPersona):
        """添加购买者画像工作表（带饼图）"""
        ws['A1'] = "购买者画像统计"
        ws['A1'].font = Font(size=16, bold=True)

        row = 3
        ws[f'A{row}'] = "平均年龄"
        ws[f'B{row}'] = f"{persona.avg_age:.1f} 岁"
        row += 1
        ws[f'A{row}'] = "平均收入"
        ws[f'B{row}'] = f"¥{persona.avg_income:,.0f}"
        row += 1
        ws[f'A{row}'] = "主要职业"
        ws[f'B{row}'] = persona.top_occupation
        row += 1
        ws[f'A{row}'] = "主要学历"
        ws[f'B{row}'] = persona.top_education
        row += 2

        ws[f'A{row}'] = "年龄分布"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        age_row = row
        for age_bucket, count in sorted(persona.age_distribution.items()):
            ws[f'A{row}'] = age_bucket
            ws[f'B{row}'] = count
            row += 1

        if row > age_row + 1:
            try:
                chart = PieChart()
                chart.title = "年龄分布"
                data_ref = Reference(ws, min_col=2, min_row=age_row, max_row=row-1)
                cats_ref = Reference(ws, min_col=1, min_row=age_row, max_row=row-1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 12
                chart.height = 10
                ws.add_chart(chart, "D3")
            except Exception as e:
                print(f"    [WARN] 年龄饼图失败：{e}")

    def _analyze_non_buyer_reasons(self, results: Dict) -> NonBuyerReason:
        """分析未购买原因"""
        non_buyers = results.get('non_buyers', [])

        price_sensitive = 0
        not_needed = 0
        has_alternative = 0
        other = 0

        for nb in non_buyers:
            reason = nb.get('reason', 'other')
            if 'price' in reason.lower() or '贵' in reason or '便宜' in reason:
                price_sensitive += 1
            elif 'need' in reason.lower() or '需要' in reason or '想要' in reason:
                not_needed += 1
            elif 'alternative' in reason.lower() or '替代' in reason or '已有' in reason:
                has_alternative += 1
            else:
                other += 1

        return NonBuyerReason(
            price_sensitive=price_sensitive,
            not_needed=not_needed,
            has_alternative=has_alternative,
            other=other,
            total=len(non_buyers)
        )

    def _add_non_buyer_analysis(self, ws, non_buyer: NonBuyerReason):
        """添加未购买分析工作表"""
        ws['A1'] = "未购买原因分析"
        ws['A1'].font = Font(size=16, bold=True)

        row = 3
        ws[f'A{row}'] = "原因"
        ws[f'B{row}'] = "人数"
        ws[f'C{row}'] = "占比"
        for col in 'ABC':
            ws[f'{col}{row}'].font = Font(bold=True)

        row += 1
        reasons = [
            ("价格敏感", non_buyer.price_sensitive, non_buyer.price_sensitive_pct),
            ("不需要", non_buyer.not_needed, non_buyer.not_needed_pct),
            ("已有替代品", non_buyer.has_alternative, non_buyer.has_alternative_pct),
            ("其他", non_buyer.other, non_buyer.other / non_buyer.total * 100 if non_buyer.total > 0 else 0),
        ]

        for reason, count, pct in reasons:
            if count > 0:
                ws[f'A{row}'] = reason
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{pct:.1f}%"
                row += 1

        if row > 4:
            try:
                chart = BarChart()
                chart.title = "未购买原因分布"
                chart.type = "col"
                data_ref = Reference(ws, min_col=2, min_row=3, max_row=row-1)
                cats_ref = Reference(ws, min_col=1, min_row=4, max_row=row-1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 15
                chart.height = 10
                ws.add_chart(chart, "E3")
            except Exception as e:
                print(f"    [WARN] 未购买原因图表失败：{e}")

    def _evaluate_success(self, results: Dict, user_config: Dict) -> Dict:
        """评估实验是否成功"""
        metrics = results.get('metrics', {})
        criteria = user_config.get('success_criteria', {})

        achieved = []
        not_achieved = []

        for metric, target in criteria.items():
            actual = metrics.get(metric, 0)
            if actual >= target:
                achieved.append(f"{metric}: {actual*100:.1f}% (目标{target*100:.0f}%)")
            else:
                not_achieved.append(f"{metric}: {actual*100:.1f}% (目标{target*100:.0f}%)")

        return {
            'overall_success': len(achieved) > len(not_achieved),
            'achieved': achieved,
            'not_achieved': not_achieved,
            'key_findings': [
                f"共{len(achieved)}项指标达标，{len(not_achieved)}项未达标",
                f"转化率：{metrics.get('conversion_rate', 0)*100:.1f}%",
                f"购买者{len(results.get('buyers', []))}人，未购买{len(results.get('non_buyers', []))}人",
            ]
        }

    def _add_recommendations(self, ws, results: Dict, user_config: Dict, persona: BuyerPersona, non_buyer: NonBuyerReason):
        """添加改进建议工作表"""
        ws['A1'] = "改进建议与下一步行动"
        ws['A1'].font = Font(size=16, bold=True)

        row = 3
        ws[f'A{row}'] = "目标用户画像优化"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"建议聚焦：{persona.top_occupation}，{persona.top_education}，"
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
        ws[f'A{row}'] = f"年龄{persona.avg_age:.0f}岁左右，月收入¥{persona.avg_income:,.0f}"
        row += 2

        ws[f'A{row}'] = "未购买原因应对"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        if non_buyer.price_sensitive > non_buyer.not_needed and non_buyer.price_sensitive > non_buyer.has_alternative:
            ws[f'A{row}'] = "主要障碍：价格敏感"
            ws[f'A{row+1}'] = "建议：考虑降价 10-15% / 推出入门版 / 增加分期付款"
        elif non_buyer.not_needed > non_buyer.price_sensitive:
            ws[f'A{row}'] = "主要障碍：需求不足"
            ws[f'A{row+1}'] = "建议：加强用户教育 / 突出核心价值 / 场景化营销"
        else:
            ws[f'A{row}'] = "主要障碍：竞争激烈"
            ws[f'A{row+1}'] = "建议：差异化定位 / 强调独特优势 / 建立品牌认知"
        row += 3

        ws[f'A{row}'] = "下一步行动"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "1. 根据画像优化目标用户定位"
        row += 1
        ws[f'A{row}'] = "2. 针对主要障碍调整策略"
        row += 1
        ws[f'A{row}'] = "3. 进行 A/B 测试验证改进效果"
        row += 1
        ws[f'A{row}'] = "4. 扩大样本量进行二次验证"

    def _generate_csv_data(self, experiment_id: str, results: Dict, timestamp: str) -> Optional[str]:
        """生成 CSV 数据文件"""
        if not self.data_exporter:
            return None

        try:
            buyers = results.get('buyers', [])
            if buyers:
                buyer_data = []
                for b in buyers:
                    buyer_data.append({
                        'id': b.get('id'),
                        'name': b.get('name'),
                        'age': b.get('age'),
                        'gender': b.get('gender'),
                        'occupation': b.get('occupation'),
                        'income': b.get('income'),
                        'education_level': b.get('education_level'),
                        'marital_status': b.get('marital_status'),
                        'health_score': b.get('health_score'),
                        'happiness': b.get('happiness'),
                    })

                csv_path = self.data_exporter.export_to_csv(
                    buyer_data,
                    f"{experiment_id}_buyers_{timestamp}.csv"
                )
                print(f"  [OK] CSV 数据：{os.path.basename(csv_path)}")
                return csv_path
        except Exception as e:
            print(f"  ❌ CSV 生成失败：{e}")

        return None

    def _generate_text_summary(self, experiment_id: str, results: Dict, user_config: Dict, timestamp: str) -> Optional[str]:
        """生成文本摘要（浏览器可查看）"""
        try:
            success = self._evaluate_success(results, user_config)
            persona = self._analyze_buyer_persona(results)
            non_buyer = self._analyze_non_buyer_reasons(results)

            summary = f"""# 实验报告摘要

**实验 ID**: {experiment_id}
**实验模板**: {user_config.get('template_name', '未知')}
**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 实验结论

{'✅ 成功' if success['overall_success'] else '❌ 未达目标'}

### 指标达成
"""
            for item in success['achieved']:
                summary += f"- ✓ {item}\n"
            for item in success['not_achieved']:
                summary += f"- ✗ {item}\n"

            summary += f"""
## 购买者画像

- **平均年龄**: {persona.avg_age:.1f} 岁
- **平均收入**: ¥{persona.avg_income:,.0f}
- **主要职业**: {persona.top_occupation}
- **主要学历**: {persona.top_education}

## 未购买原因

- 价格敏感：{non_buyer.price_sensitive}人 ({non_buyer.price_sensitive_pct:.1f}%)
- 不需要：{non_buyer.not_needed}人 ({non_buyer.not_needed_pct:.1f}%)
- 已有替代品：{non_buyer.has_alternative}人 ({non_buyer.has_alternative_pct:.1f}%)
- 其他：{non_buyer.other}人 ({non_buyer.other / non_buyer.total * 100 if non_buyer.total > 0 else 0:.1f}%)

## 关键发现

"""
            for finding in success['key_findings']:
                summary += f"- {finding}\n"

            filename = f"summary_{experiment_id}_{timestamp}.md"
            filepath = os.path.join(self.output_dir, filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(summary)

            print(f"  [OK] 文本摘要：{filename}")
            return filepath

        except Exception as e:
            print(f"  ❌ 文本摘要失败：{e}")
            return None
