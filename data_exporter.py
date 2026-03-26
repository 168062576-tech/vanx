"""
御龙军虚拟 Agent 世界 - 数据导出模块
版本：v2.1 (移植自旧版 v2.0)
创建时间：2026-03-16
移植时间：2026-03-24
功能：导出 CSV/Excel/JSON 格式数据
"""

import sys
import json
import csv
import os
from datetime import datetime
from typing import List, Dict


class DataExporter:
    """数据导出器"""

    def __init__(self, output_dir: str = 'exports'):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def export_to_csv(self, data: List[Dict], filename: str) -> str:
        """
        导出为 CSV 格式

        Args:
            data: 数据列表
            filename: 文件名

        Returns:
            文件路径
        """
        filepath = os.path.join(self.output_dir, filename)

        if not data:
            print("[警告] 数据为空，跳过导出")
            return filepath

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

        print(f"[OK] CSV 导出完成：{filepath}")
        return filepath

    def export_to_excel(self, data: List[Dict], filename: str) -> str:
        """
        导出为 Excel 格式

        Args:
            data: 数据列表
            filename: 文件名

        Returns:
            文件路径
        """
        try:
            import openpyxl
            from openpyxl import Workbook

            filepath = os.path.join(self.output_dir, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = '数据'

            # 写入表头
            headers = list(data[0].keys()) if data else []
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入数据
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data.values(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(filepath)
            print(f"[OK] Excel 导出完成：{filepath}")
            return filepath

        except ImportError:
            print("[警告] openpyxl 未安装，降级为 CSV 导出")
            return self.export_to_csv(data, filename.replace('.xlsx', '.csv'))

    def export_to_json(self, data: Dict, filename: str) -> str:
        """
        导出为 JSON 格式

        Args:
            data: 数据字典
            filename: 文件名

        Returns:
            文件路径
        """
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"[OK] JSON 导出完成：{filepath}")
        return filepath

    def export_agents(self, agents: List[Dict], format: str = 'csv') -> str:
        """
        导出 Agent 数据

        Args:
            agents: Agent 列表
            format: 导出格式 ('csv', 'excel', 'json')

        Returns:
            文件路径
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if format == 'csv':
            return self.export_to_csv(agents, f'agents_{timestamp}.csv')
        elif format == 'excel':
            return self.export_to_excel(agents, f'agents_{timestamp}.xlsx')
        else:
            return self.export_to_json({'agents': agents}, f'agents_{timestamp}.json')

    def export_results(self, results: Dict, format: str = 'json') -> str:
        """
        导出演化结果

        Args:
            results: 结果字典
            format: 导出格式

        Returns:
            文件路径
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if format == 'csv':
            # 扁平化结果
            flat_data = []
            for day, metrics in results.get('daily_metrics', {}).items():
                row = {'day': day}
                row.update(metrics)
                flat_data.append(row)
            return self.export_to_csv(flat_data, f'results_{timestamp}.csv')
        elif format == 'excel':
            flat_data = []
            for day, metrics in results.get('daily_metrics', {}).items():
                row = {'day': day}
                row.update(metrics)
                flat_data.append(row)
            return self.export_to_excel(flat_data, f'results_{timestamp}.xlsx')
        else:
            return self.export_to_json(results, f'results_{timestamp}.json')


# 使用示例
if __name__ == "__main__":
    print("=" * 60)
    print("数据导出模块 - 使用示例")
    print("=" * 60)
    print()

    exporter = DataExporter()

    # 测试 Agent 数据
    test_agents = [
        {'id': 'agent_001', 'occupation': 'IT', 'satisfaction': 4.5, 'stress': 2.3},
        {'id': 'agent_002', 'occupation': '金融', 'satisfaction': 3.8, 'stress': 5.6},
        {'id': 'agent_003', 'occupation': '医疗', 'satisfaction': 4.2, 'stress': 3.1}
    ]

    # 导出 CSV
    print("1. 导出 CSV:")
    csv_file = exporter.export_agents(test_agents, 'csv')
    print()

    # 导出 Excel
    print("2. 导出 Excel:")
    excel_file = exporter.export_agents(test_agents, 'excel')
    print()

    # 导出 JSON
    print("3. 导出 JSON:")
    json_file = exporter.export_agents(test_agents, 'json')
    print()

    print("[OK] 数据导出模块测试完成！")
